#!/usr/bin/env python3
"""
Standalone car review analysis script.

Edit the CONFIG section below, then run:
    python car_review_analyzer_standalone.py
"""

from __future__ import annotations

import json
import re
import statistics
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openai import OpenAI
from openpyxl import Workbook, load_workbook


# =========================
# CONFIG
# =========================
INPUT_EXCEL = "reviews.xlsx"
OUTPUT_MARKDOWN = "review_analysis_report.md"
OUTPUT_SUMMARY_EXCEL = "review_analysis_summary.xlsx"

# auto: if API key is configured, try LLM first, then fall back to heuristic
# llm: force LLM analysis
# heuristic: force local heuristic analysis
ANALYSIS_MODE = "auto"

API_BASE = "https://aihubmix.com/v1"
API_KEY = "PUT_YOUR_API_KEY_HERE"
MODEL_NAME = "Qwen/Qwen2.5-7B-Instruct"
MAX_COMMENTS_FOR_LLM = 80


ASPECTS = {
    "外观": {
        "keywords": ["外观", "颜值", "帅", "好看", "回头率", "造型", "设计", "低趴", "车漆"],
        "topics": ["颜值高", "外观好看", "设计感", "回头率", "低趴", "车漆", "轮毂", "前脸", "尾部"],
        "positive": ["好看", "帅", "喜欢", "满意", "拉风", "惊艳", "漂亮", "回头率高"],
        "negative": ["普通", "一般", "不耐看", "单调", "审美疲劳"],
    },
    "内饰": {
        "keywords": ["内饰", "质感", "用料", "做工", "真皮", "塑料感", "异味", "氛围灯", "座舱"],
        "topics": ["内饰质感", "塑料感", "用料", "做工", "真皮", "异味", "高级感", "氛围灯"],
        "positive": ["精致", "高级", "有质感", "顺眼", "满意", "不错"],
        "negative": ["塑料感", "廉价", "一般", "异味", "粗糙"],
    },
    "空间": {
        "keywords": ["空间", "后排", "前排", "头部", "腿部", "后备箱", "储物", "顶头", "局促"],
        "topics": ["后排空间", "前排空间", "头部空间", "腿部空间", "后备箱", "顶头", "局促", "储物空间"],
        "positive": ["宽敞", "够用", "大", "充足", "满意"],
        "negative": ["顶头", "局促", "小", "拥挤", "不够"],
    },
    "舒适性": {
        "keywords": ["舒适", "座椅", "隔音", "噪音", "异响", "颠簸", "滤震", "风噪", "胎噪"],
        "topics": ["座椅偏硬", "隔音", "异响", "腰酸背痛", "风噪", "胎噪", "滤震", "底盘舒适"],
        "positive": ["舒服", "舒适", "安静", "平稳", "满意"],
        "negative": ["偏硬", "太硬", "异响", "噪音", "风噪", "胎噪", "腰酸背痛"],
    },
    "动力操控": {
        "keywords": ["动力", "加速", "操控", "转向", "底盘", "过弯", "刹车", "超车", "驾驶感受"],
        "topics": ["加速快", "转向精准", "底盘调校", "过弯", "超车", "刹车", "操控灵活", "打滑"],
        "positive": ["灵活", "精准", "稳", "丝滑", "给力", "强", "满意"],
        "negative": ["打滑", "偏软", "偏硬", "不足", "发飘"],
    },
    "续航补能": {
        "keywords": ["续航", "能耗", "电耗", "充电", "快充", "慢充", "补能", "充电桩", "公里"],
        "topics": ["续航扎实", "电耗", "快充", "慢充", "补能效率", "续航焦虑", "充电速度", "充电桩"],
        "positive": ["扎实", "省电", "够用", "满意", "优秀", "快"],
        "negative": ["费电", "慢", "虚", "焦虑", "掉电快", "不准"],
    },
    "智能化": {
        "keywords": ["车机", "语音", "导航", "智驾", "辅助驾驶", "自动泊车", "OTA", "小爱", "APP"],
        "topics": ["车机流畅", "语音控制", "智驾", "辅助驾驶", "自动泊车", "OTA", "卡顿", "误报"],
        "positive": ["流畅", "好用", "聪明", "方便", "稳定", "满意"],
        "negative": ["卡顿", "误报", "不好用", "不稳定", "保守", "激进"],
    },
    "性价比": {
        "keywords": ["价格", "性价比", "配置", "保险", "花费", "预算", "落地", "选配", "值"],
        "topics": ["性价比", "价格", "配置", "保险", "选配", "落地价", "花费", "偏贵"],
        "positive": ["值", "划算", "厚道", "实在", "满意"],
        "negative": ["贵", "偏贵", "不值", "鸡肋", "溢价"],
    },
}

GENERAL_POSITIVE_WORDS = [
    "满意", "喜欢", "优秀", "不错", "好用", "稳定", "顺滑", "实用", "惊喜", "舒服", "省心", "给力", "靠谱",
]
GENERAL_NEGATIVE_WORDS = [
    "不满意", "一般", "差", "问题", "异响", "顶头", "费电", "太硬", "不足", "失望", "卡顿", "难受", "偏贵",
]
NEGATIVE_HARD_WORDS = {
    "顶头": 1.0,
    "局促": 0.9,
    "异响": 1.0,
    "太硬": 1.0,
    "偏硬": 0.9,
    "卡顿": 0.9,
    "误报": 0.9,
    "费电": 0.9,
    "慢": 0.5,
    "贵": 0.6,
    "塑料感": 0.7,
    "焦虑": 0.8,
}
STOPWORDS = {
    "一个", "这个", "那个", "我们", "你们", "他们", "真的", "觉得", "还是", "就是", "比较", "非常", "而且", "因为",
    "如果", "已经", "没有", "可以", "一些", "不是", "自己", "时候", "感觉", "目前", "现在", "整体", "方面", "车型",
    "这台", "这车", "车辆", "小米", "标准版", "pro", "max", "su7", "一下", "这样", "还有", "然后", "以后",
}


@dataclass
class Review:
    review_time: str
    car_model: str
    rating: float | None
    content: str


@dataclass
class AnalysisResult:
    sheet_name: str
    car_model: str
    review_count: int
    user_avg_rating: float
    model_score: float
    overall_score: float
    summary: str
    strengths: List[str]
    weaknesses: List[str]
    weakness_details: List[str]
    purchase_advice: str
    aspect_scores: Dict[str, float]
    aspect_analysis: Dict[str, str]
    aspect_keywords: Dict[str, str]
    aspect_negative_ratio: Dict[str, float]
    method: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def safe_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        match = re.search(r"(\d+(?:\.\d+)?)", str(value))
        return float(match.group(1)) if match else None


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def split_sentences(text: str) -> List[str]:
    parts = re.split(r"[。！？；!?]+", clean_text(text))
    return [part.strip() for part in parts if len(part.strip()) >= 4]


def load_reviews_by_sheet(path: str) -> Dict[str, List[Review]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: Dict[str, List[Review]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            continue
        reviews: List[Review] = []
        for row in rows[1:]:
            if not row:
                continue
            review_time = clean_text(row[0] if len(row) > 0 else "")
            car_model = clean_text(row[1] if len(row) > 1 else "")
            rating = safe_float(row[2] if len(row) > 2 else "")
            content = clean_text(row[3] if len(row) > 3 else "")
            if content:
                reviews.append(Review(review_time, car_model, rating, content))
        if reviews:
            result[sheet_name] = reviews
    return result


def average_rating(reviews: List[Review]) -> float:
    ratings = [review.rating for review in reviews if review.rating is not None]
    return round(statistics.mean(ratings), 2) if ratings else 0.0


def count_hits(text: str, words: List[str]) -> int:
    return sum(text.count(word) for word in words)


def extract_tokens(text: str) -> List[str]:
    chunks = re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,8}", text)
    return [token for token in chunks if token not in STOPWORDS]


def extract_topic_keywords(sentences: List[str], aspect_spec: Dict[str, List[str]]) -> List[str]:
    combined = " ".join(sentences)
    topic_counter = Counter()
    for topic in aspect_spec.get("topics", []):
        hits = combined.count(topic)
        if hits:
            topic_counter[topic] += hits
    if topic_counter:
        return [topic for topic, _ in topic_counter.most_common(3)]

    token_counter = Counter()
    for sentence in sentences:
        token_counter.update(extract_tokens(sentence))
    for token in (
        aspect_spec.get("keywords", [])
        + aspect_spec.get("topics", [])
        + aspect_spec.get("positive", [])
        + aspect_spec.get("negative", [])
        + GENERAL_POSITIVE_WORDS
        + GENERAL_NEGATIVE_WORDS
    ):
        token_counter.pop(token, None)
    return [token for token, _ in token_counter.most_common(3)]


def representative_sentence(sentences: List[str], positive_words: List[str], negative_words: List[str], prefer_negative: bool) -> str:
    preferred = negative_words if prefer_negative else positive_words
    backup = positive_words + negative_words if prefer_negative else negative_words + positive_words
    for word in preferred:
        for sentence in sentences:
            if word in sentence and 8 <= len(sentence) <= 90:
                return sentence
    for word in backup:
        for sentence in sentences:
            if word in sentence and 8 <= len(sentence) <= 90:
                return sentence
    for sentence in sentences:
        if 8 <= len(sentence) <= 90:
            return sentence
    return ""


def sentence_sentiment(sentence: str, positive_words: List[str], negative_words: List[str]) -> Tuple[float, float, float]:
    pos = count_hits(sentence, positive_words + GENERAL_POSITIVE_WORDS)
    neg = count_hits(sentence, negative_words + GENERAL_NEGATIVE_WORDS)
    hard_neg = sum(weight for word, weight in NEGATIVE_HARD_WORDS.items() if word in sentence)
    if pos == 0 and neg == 0 and hard_neg == 0:
        return 0.0, 0.0, 0.0
    sentiment = pos * 0.8 - neg * 0.9 - hard_neg
    return sentiment, float(pos), float(neg + hard_neg)


def score_from_text(content: str) -> float:
    pos = count_hits(content, GENERAL_POSITIVE_WORDS)
    neg = count_hits(content, GENERAL_NEGATIVE_WORDS)
    raw = 3.2 + min(1.1, pos * 0.07) - min(1.2, neg * 0.11)
    return round(clamp(raw, 1.0, 5.0), 2)


def weakness_label(aspect: str, negative_keyword: str, aspect_keywords: List[str]) -> str:
    if negative_keyword:
        return f"{aspect}短板主要集中在“{negative_keyword}”"
    if aspect_keywords:
        return f"{aspect}短板主要集中在“{aspect_keywords[0]}”"
    return f"{aspect}存在需要重点确认的短板"


def summarize_aspects_heuristic(reviews: List[Review]):
    aspect_scores: Dict[str, float] = {}
    aspect_analysis: Dict[str, str] = {}
    aspect_keywords: Dict[str, str] = {}
    aspect_negative_ratio: Dict[str, float] = {}
    strengths: List[str] = []
    weaknesses: List[str] = []
    weakness_details: List[str] = []
    weakness_candidates: List[Tuple[float, str, str]] = []

    explicit_ratings = [review.rating for review in reviews if review.rating is not None]
    rating_anchor = statistics.mean(explicit_ratings) if explicit_ratings else 3.8

    for aspect, spec in ASPECTS.items():
        keywords = spec["keywords"]
        positive_words = spec["positive"]
        negative_words = spec["negative"]

        matched_reviews = [review for review in reviews if any(keyword in review.content for keyword in keywords)]
        if not matched_reviews:
            aspect_scores[aspect] = 0.0
            aspect_analysis[aspect] = "评论中关于该维度的信息较少，暂时无法形成稳定判断。"
            aspect_keywords[aspect] = "关键词不足"
            aspect_negative_ratio[aspect] = 0.0
            continue

        matched_sentences = []
        for review in matched_reviews:
            review_sentences = split_sentences(review.content)
            relevant = [sentence for sentence in review_sentences if any(keyword in sentence for keyword in keywords)]
            matched_sentences.extend(relevant or review_sentences[:1])

        sentence_scores = [sentence_sentiment(sentence, positive_words, negative_words) for sentence in matched_sentences]
        informative_scores = [item for item in sentence_scores if any(item)]
        if not informative_scores:
            informative_scores = [(0.0, 0.0, 0.0)]

        avg_sentiment = statistics.mean(item[0] for item in informative_scores)
        positive_sentence_ratio = sum(1 for item in informative_scores if item[0] > 0.15) / len(informative_scores)
        negative_sentence_ratio = sum(1 for item in informative_scores if item[0] < -0.15) / len(informative_scores)
        mention_ratio = len(matched_reviews) / max(1, len(reviews))

        base_score = 3.0 + avg_sentiment * 0.65
        rating_bonus = ((rating_anchor - 3.5) / 1.5) * 0.45
        mention_bonus = min(0.25, mention_ratio * 0.25)
        negative_penalty = negative_sentence_ratio * 1.0
        positive_bonus = positive_sentence_ratio * 0.35
        aspect_score = round(clamp(base_score + rating_bonus + mention_bonus + positive_bonus - negative_penalty, 1.2, 4.8), 2)

        combined_text = " ".join(matched_sentences)
        pos_hits = count_hits(combined_text, positive_words + GENERAL_POSITIVE_WORDS)
        neg_hits = count_hits(combined_text, negative_words + GENERAL_NEGATIVE_WORDS)
        top_keywords = extract_topic_keywords(matched_sentences, spec) or keywords[:3]
        positive_keyword = next((word for word in positive_words + GENERAL_POSITIVE_WORDS if word in combined_text), "")
        negative_keyword = next((word for word in negative_words + GENERAL_NEGATIVE_WORDS if word in combined_text), "")
        positive_sample = representative_sentence(matched_sentences, positive_words + GENERAL_POSITIVE_WORDS, negative_words + GENERAL_NEGATIVE_WORDS, False)
        negative_sample = representative_sentence(matched_sentences, positive_words + GENERAL_POSITIVE_WORDS, negative_words + GENERAL_NEGATIVE_WORDS, True)

        aspect_scores[aspect] = aspect_score
        aspect_keywords[aspect] = "、".join(top_keywords)
        aspect_negative_ratio[aspect] = round(negative_sentence_ratio, 2)

        analysis_parts = [
            f"命中评论{len(matched_reviews)}条，占比{mention_ratio:.0%}",
            f"负向句子占比约{negative_sentence_ratio:.0%}",
        ]
        if positive_keyword:
            analysis_parts.append(f"正向关键词偏向“{positive_keyword}”")
        if negative_keyword:
            analysis_parts.append(f"负向关键词偏向“{negative_keyword}”")
        if top_keywords:
            analysis_parts.append(f"高频主题词包括“{'、'.join(top_keywords)}”")
        if positive_sample and negative_sentence_ratio < 0.45:
            analysis_parts.append(f"代表性反馈如“{positive_sample}”")
        elif negative_sample:
            analysis_parts.append(f"代表性反馈如“{negative_sample}”")
        aspect_analysis[aspect] = "；".join(analysis_parts) + "。"

        if aspect_score >= 4.05 and negative_sentence_ratio <= 0.3:
            strengths.append(f"{aspect}是稳定优势项，正向反馈更集中")

        weakness_trigger = (
            aspect_score <= 3.45
            or negative_sentence_ratio >= 0.42
            or (negative_keyword in {"顶头", "异响", "太硬", "偏硬", "误报", "卡顿"} and negative_sentence_ratio >= 0.25)
        )
        if weakness_trigger:
            explanation = [
                f"{aspect}得分为{aspect_score:.2f}/5",
                f"命中评论{len(matched_reviews)}条，占比{mention_ratio:.0%}",
                f"负向句子占比约{negative_sentence_ratio:.0%}",
            ]
            if negative_keyword:
                explanation.append(f"负向关键词集中在“{negative_keyword}”")
            if top_keywords:
                explanation.append(f"高频主题词包括“{'、'.join(top_keywords)}”")
            if negative_sample:
                explanation.append(f"代表性差评如“{negative_sample}”")
            weakness_candidates.append(
                (
                    aspect_score,
                    weakness_label(aspect, negative_keyword, top_keywords),
                    "；".join(explanation) + "。",
                )
            )

    strengths = strengths[:4]
    if not strengths:
        ranked = sorted([(aspect, score) for aspect, score in aspect_scores.items() if score > 0], key=lambda item: item[1], reverse=True)
        for aspect, score in ranked[:2]:
            if score >= 3.8:
                strengths.append(f"{aspect}整体表现较好，是用户认可度更高的方面")
    strengths = strengths[:4] or ["整体体验较均衡，优势更集中在外观、配置或智能体验等方面。"]

    weakness_candidates.sort(key=lambda item: item[0])
    for _, label, detail in weakness_candidates:
        if label not in weaknesses:
            weaknesses.append(label)
            weakness_details.append(detail)

    if not weaknesses:
        ranked = sorted([(aspect, score) for aspect, score in aspect_scores.items() if score > 0], key=lambda item: item[1])
        for aspect, score in ranked[:2]:
            if score <= 3.8:
                label = weakness_label(aspect, "", aspect_keywords.get(aspect, "").split("、"))
                weaknesses.append(label)
                weakness_details.append(
                    f"{aspect}是当前相对弱项；得分为{score:.2f}/5；关键词包括“{aspect_keywords.get(aspect, '无')}”；"
                    f"建议线下重点核验这一维度。"
                )
                if len(weaknesses) >= 2:
                    break

    weaknesses = weaknesses[:3]
    weakness_details = weakness_details[:3]
    return aspect_scores, aspect_analysis, aspect_keywords, aspect_negative_ratio, strengths, weaknesses, weakness_details


def heuristic_analysis(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    car_model = next((review.car_model for review in reviews if review.car_model), sheet_name)
    user_avg = average_rating(reviews)
    text_scores = [score_from_text(review.content) for review in reviews]
    model_score = round(statistics.mean(text_scores), 2) if text_scores else user_avg
    overall = round(user_avg * 0.7 + model_score * 0.3, 2) if user_avg else model_score

    (
        aspect_scores,
        aspect_analysis,
        aspect_keywords,
        aspect_negative_ratio,
        strengths,
        weaknesses,
        weakness_details,
    ) = summarize_aspects_heuristic(reviews)

    ranked = sorted([(aspect, score) for aspect, score in aspect_scores.items() if score > 0], key=lambda item: item[1], reverse=True)
    best_aspect = ranked[0][0] if ranked else "整体表现"
    weakest_aspect = ranked[-1][0] if ranked else "短板项"
    summary = (
        f"{car_model}共分析{len(reviews)}条评论。用户平均评分为{user_avg:.2f}/5，"
        f"文本分析分为{model_score:.2f}/5，综合评分为{overall:.2f}/5。"
        f"从评论反馈看，{best_aspect}更容易获得正面认可，{weakest_aspect}是更需要重点确认的维度。"
    )
    purchase_advice = (
        f"如果你最看重{best_aspect}，这款车整体值得优先试驾；"
        f"如果你对{weakest_aspect}敏感，建议在线下把这一项验证清楚再决定。"
    )

    return AnalysisResult(
        sheet_name=sheet_name,
        car_model=car_model,
        review_count=len(reviews),
        user_avg_rating=user_avg,
        model_score=model_score,
        overall_score=overall,
        summary=summary,
        strengths=strengths,
        weaknesses=weaknesses,
        weakness_details=weakness_details,
        purchase_advice=purchase_advice,
        aspect_scores=aspect_scores,
        aspect_analysis=aspect_analysis,
        aspect_keywords=aspect_keywords,
        aspect_negative_ratio=aspect_negative_ratio,
        method="heuristic",
    )


def build_llm_payload(reviews: List[Review], max_comments: int) -> List[Dict[str, str]]:
    payload = []
    for index, review in enumerate(reviews[:max_comments], start=1):
        payload.append(
            {
                "index": str(index),
                "time": review.review_time,
                "car_model": review.car_model,
                "rating": "" if review.rating is None else f"{review.rating:.2f}",
                "content": review.content[:1200],
            }
        )
    return payload


def extract_json_from_text(text: str) -> Dict:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        text = text[start : end + 1]
    return json.loads(text)


def llm_analysis(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    car_model = next((review.car_model for review in reviews if review.car_model), sheet_name)
    user_avg = average_rating(reviews)
    client = OpenAI(api_key=API_KEY, base_url=API_BASE)
    prompt = {
        "car_model": car_model,
        "review_count": len(reviews),
        "user_average_rating": user_avg,
        "task": (
            "你是一名汽车评论分析师。请基于用户购后评论输出JSON。"
            "要求输出summary、model_score(1-5)、strengths、weaknesses、weakness_details、purchase_advice、"
            "aspect_scores、aspect_analysis、aspect_keywords、aspect_negative_ratio。"
            "优点要保留真实好的方面，短板只保留少量但真实、有依据的问题。"
            "不要为了找短板而夸大问题，也不要因为整车评分高就掩盖明显短板。"
        ),
        "aspects": list(ASPECTS.keys()),
        "comments": build_llm_payload(reviews, MAX_COMMENTS_FOR_LLM),
    }
    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": "请只返回合法JSON，不要输出代码块，不要解释。"},
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        temperature=0.2,
    )
    data = extract_json_from_text(response.choices[0].message.content.strip())

    aspect_scores = {aspect: round(float(data.get("aspect_scores", {}).get(aspect, 0)), 2) for aspect in ASPECTS}
    aspect_analysis = {aspect: clean_text(data.get("aspect_analysis", {}).get(aspect, "")) for aspect in ASPECTS}
    aspect_keywords = {aspect: clean_text(data.get("aspect_keywords", {}).get(aspect, "")) for aspect in ASPECTS}
    aspect_negative_ratio = {aspect: round(float(data.get("aspect_negative_ratio", {}).get(aspect, 0)), 2) for aspect in ASPECTS}
    model_score = round(float(data.get("model_score", user_avg or 3.5)), 2)
    overall = round(user_avg * 0.7 + model_score * 0.3, 2) if user_avg else model_score

    return AnalysisResult(
        sheet_name=sheet_name,
        car_model=car_model,
        review_count=len(reviews),
        user_avg_rating=user_avg,
        model_score=model_score,
        overall_score=overall,
        summary=clean_text(data.get("summary", "")),
        strengths=[clean_text(item) for item in data.get("strengths", [])][:4],
        weaknesses=[clean_text(item) for item in data.get("weaknesses", [])][:3],
        weakness_details=[clean_text(item) for item in data.get("weakness_details", [])][:3],
        purchase_advice=clean_text(data.get("purchase_advice", "")),
        aspect_scores=aspect_scores,
        aspect_analysis=aspect_analysis,
        aspect_keywords=aspect_keywords,
        aspect_negative_ratio=aspect_negative_ratio,
        method=f"llm:{MODEL_NAME}",
    )


def analyze_sheet(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    if ANALYSIS_MODE == "heuristic":
        return heuristic_analysis(sheet_name, reviews)
    if ANALYSIS_MODE == "llm":
        return llm_analysis(sheet_name, reviews)
    if API_KEY and API_KEY != "PUT_YOUR_API_KEY_HERE":
        try:
            return llm_analysis(sheet_name, reviews)
        except Exception:
            return heuristic_analysis(sheet_name, reviews)
    return heuristic_analysis(sheet_name, reviews)


def render_markdown(results: List[AnalysisResult]) -> str:
    lines: List[str] = ["汽车评论分析报告", ""]
    lines.append("一.总览")
    lines.append("")
    for index, item in enumerate(sorted(results, key=lambda x: x.overall_score, reverse=True), start=1):
        lines.append(
            f"{index}. {item.car_model}: 综合 {item.overall_score:.2f}/5，"
            f"用户评分均值 {item.user_avg_rating:.2f}/5，模型分析分 {item.model_score:.2f}/5，评论数 {item.review_count}"
        )
    lines.append("")

    for item in results:
        lines.append(f"二.{item.car_model}")
        lines.append("")
        lines.append(f"· 分析方式：`{item.method}`")
        lines.append(f"· 评论数：`{item.review_count}`")
        lines.append(f"· 用户平均评分：`{item.user_avg_rating:.2f}/5`")
        lines.append(f"· 模型分析分：`{item.model_score:.2f}/5`")
        lines.append(f"· 综合评分：`{item.overall_score:.2f}/5`")
        lines.append("")
        lines.append("三.模型分析结果")
        lines.append(item.summary)
        lines.append("")
        lines.append("四.主要优点")
        for strength in item.strengths:
            lines.append(f"· {strength}")
        lines.append("")
        lines.append("五.主要短板")
        for weakness in item.weaknesses:
            lines.append(f"· {weakness}")
        lines.append("")
        if item.weakness_details:
            lines.append("六.短板解释")
            for detail in item.weakness_details:
                lines.append(f"· {detail}")
                lines.append("")
            lines.append("")
        lines.append("七.用户各方面体验分析")
        for aspect in ASPECTS:
            lines.append(
                f"· {aspect}: {item.aspect_scores.get(aspect, 0.0):.2f}/5。"
                f"负向句子占比：{item.aspect_negative_ratio.get(aspect, 0.0):.0%}。"
                f"关键词：{item.aspect_keywords.get(aspect, '无')}。"
                f"{item.aspect_analysis.get(aspect, '')}"
            )
            lines.append("")  # 【修改点】每个分析点后增加空行，避免拥挤
        lines.append("")
        lines.append("八.建议")
        lines.append(item.purchase_advice)
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "车型分析"
    return cleaned[:31]


def write_summary_excel(path: str, results: List[AnalysisResult]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "分析汇总"
    summary.append(["车型", "评论数", "用户平均评分", "模型分析分", "综合评分(5分制)", "分析方式", "建议"])

    for item in sorted(results, key=lambda x: x.overall_score, reverse=True):
        summary.append(
            [
                item.car_model,
                item.review_count,
                item.user_avg_rating,
                item.model_score,
                item.overall_score,
                item.method,
                item.purchase_advice,
            ]
        )

    for item in results:
        sheet = workbook.create_sheet(safe_sheet_name(item.car_model))
        sheet.append(["维度", "得分(5分制)", "负向句子占比", "关键词", "分析", "是否主要短板", "短板解释"])
        weakness_map = {w: d for w, d in zip(item.weaknesses, item.weakness_details)}
        for aspect in ASPECTS:
            related_weakness = next((w for w in item.weaknesses if aspect in w), "")
            sheet.append(
                [
                    aspect,
                    item.aspect_scores.get(aspect, 0.0),
                    item.aspect_negative_ratio.get(aspect, 0.0),
                    item.aspect_keywords.get(aspect, ""),
                    item.aspect_analysis.get(aspect, ""),
                    "是" if related_weakness else "否",
                    weakness_map.get(related_weakness, ""),
                ]
            )

    workbook.save(path)


def write_text(path: str, text: str) -> None:
    Path(path).write_text(text, encoding="utf-8")


def main() -> None:
    if not Path(INPUT_EXCEL).exists():
        raise SystemExit(f"Input Excel not found: {INPUT_EXCEL}")
    reviews_by_sheet = load_reviews_by_sheet(INPUT_EXCEL)
    if not reviews_by_sheet:
        raise SystemExit("No review data found in the input Excel.")

    results: List[AnalysisResult] = []
    for sheet_name, reviews in reviews_by_sheet.items():
        results.append(analyze_sheet(sheet_name, reviews))

    write_text(OUTPUT_MARKDOWN, render_markdown(results))
    write_summary_excel(OUTPUT_SUMMARY_EXCEL, results)

    print(f"input={INPUT_EXCEL}")
    print(f"sheets={len(reviews_by_sheet)}")
    print(f"report={OUTPUT_MARKDOWN}")
    print(f"summary_excel={OUTPUT_SUMMARY_EXCEL}")


if __name__ == "__main__":
    main()
