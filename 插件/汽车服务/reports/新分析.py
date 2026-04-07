#!/usr/bin/env python3
"""
Standalone car review analysis script.

Edit the CONFIG section below, then run:
    python car_review_analyzer_standalone.py
"""

from __future__ import annotations

import json
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

from openai import OpenAI
from openpyxl import Workbook, load_workbook


# =========================
# CONFIG
# =========================
INPUT_EXCEL = r"E:\2026 计算机设计大赛\llm\汽车服务\爬取评论\s.xlsx"
OUTPUT_MARKDOWN = r"E:\2026 计算机设计大赛\llm\汽车服务\分析\review_analysis_report.md"
OUTPUT_SUMMARY_EXCEL = r"E:\2026 计算机设计大赛\llm\汽车服务\分析\review_analysis_summary.xlsx"

# auto: 有可用 API Key 就走 LLM，没有就走本地规则分析
# llm: 强制走大模型分析
# heuristic: 强制走本地规则分析
ANALYSIS_MODE = "auto"

API_BASE = "https://aihubmix.com/v1"
API_KEY = "PUT_YOUR_API_KEY_HERE"
MODEL_NAME = "Qwen/Qwen2.5-7B-Instruct"

# 每个车型最多送给模型分析的评论条数
MAX_COMMENTS_FOR_LLM = 80


ASPECTS = {
    "外观": ["外观", "颜值", "帅", "好看", "回头率", "造型", "设计"],
    "内饰": ["内饰", "质感", "用料", "氛围灯", "做工"],
    "空间": ["空间", "后排", "头部", "腿部", "后备箱", "储物"],
    "舒适性": ["舒适", "座椅", "隔音", "噪音", "异响", "颠簸"],
    "动力操控": ["动力", "加速", "操控", "转向", "底盘", "过弯", "刹车"],
    "续航补能": ["续航", "能耗", "充电", "补能", "电耗", "快充"],
    "智能化": ["车机", "语音", "导航", "智驾", "辅助驾驶", "OTA", "小爱"],
    "性价比": ["价格", "性价比", "配置", "保险", "花费", "值", "预算"],
}

POSITIVE_WORDS = [
    "满意", "喜欢", "优秀", "不错", "好用", "稳定", "顺滑", "实用", "惊喜", "舒服", "省心", "给力", "靠谱",
]
NEGATIVE_WORDS = [
    "不满意", "一般", "差", "问题", "异响", "顶头", "费电", "太硬", "不足", "失望", "不稳", "卡顿", "难受", "偏贵",
]


@dataclass
class Review:
    review_time: str
    car_model: str
    rating: float | None
    content: str


@dataclass
class AnalysisResult:
    sheet_name: str
    car_model: str
    review_count: int
    user_avg_rating: float
    model_score: float
    overall_score: float
    summary: str
    strengths: List[str]
    weaknesses: List[str]
    purchase_advice: str
    aspect_scores: Dict[str, float]
    aspect_analysis: Dict[str, str]
    method: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def safe_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        match = re.search(r"(\d+(?:\.\d+)?)", str(value))
        return float(match.group(1)) if match else None


def load_reviews_by_sheet(path: str) -> Dict[str, List[Review]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: Dict[str, List[Review]] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            continue

        reviews: List[Review] = []
        for row in rows[1:]:
            if not row:
                continue
            review_time = clean_text(row[0] if len(row) > 0 else "")
            car_model = clean_text(row[1] if len(row) > 1 else "")
            rating = safe_float(row[2] if len(row) > 2 else "")
            content = clean_text(row[3] if len(row) > 3 else "")
            if content:
                reviews.append(Review(review_time, car_model, rating, content))

        if reviews:
            result[sheet_name] = reviews
    return result


def average_rating(reviews: List[Review]) -> float:
    ratings = [review.rating for review in reviews if review.rating is not None]
    return round(statistics.mean(ratings), 2) if ratings else 0.0


def score_from_text(content: str) -> float:
    pos = sum(content.count(word) for word in POSITIVE_WORDS)
    neg = sum(content.count(word) for word in NEGATIVE_WORDS)
    raw = 3.0 + min(1.6, pos * 0.08) - min(1.6, neg * 0.12)
    return round(max(1.0, min(5.0, raw)), 2)


def summarize_aspects_heuristic(reviews: List[Review]):
    aspect_scores: Dict[str, float] = {}
    aspect_analysis: Dict[str, str] = {}
    strengths: List[str] = []
    weaknesses: List[str] = []

    for aspect, keywords in ASPECTS.items():
        matched_comments = [review.content for review in reviews if any(keyword in review.content for keyword in keywords)]
        if not matched_comments:
            aspect_scores[aspect] = 0.0
            aspect_analysis[aspect] = "评论中关于该维度的信息较少。"
            continue

        joined = " ".join(matched_comments)
        score = score_from_text(joined)
        aspect_scores[aspect] = score

        positive_hits = [word for word in POSITIVE_WORDS if word in joined]
        negative_hits = [word for word in NEGATIVE_WORDS if word in joined]

        if score >= 3.8:
            strengths.append(f"{aspect}整体表现较强")
        elif score <= 2.8:
            weaknesses.append(f"{aspect}是用户集中吐槽点")

        if positive_hits and negative_hits:
            aspect_analysis[aspect] = f"{aspect}评价有分化，正面集中在“{positive_hits[0]}”，负面集中在“{negative_hits[0]}”。"
        elif positive_hits:
            aspect_analysis[aspect] = f"{aspect}整体偏正向，评论中高频提到“{positive_hits[0]}”。"
        elif negative_hits:
            aspect_analysis[aspect] = f"{aspect}偏弱，评论中高频提到“{negative_hits[0]}”。"
        else:
            aspect_analysis[aspect] = f"{aspect}有一定讨论度，但情绪倾向不够集中。"

    if not strengths:
        strengths.append("整体体验比较均衡，没有极端单项优势。")
    if not weaknesses:
        weaknesses.append("当前评论中没有非常集中的致命短板。")

    return aspect_scores, aspect_analysis, strengths[:4], weaknesses[:4]


def heuristic_analysis(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    car_model = next((review.car_model for review in reviews if review.car_model), sheet_name)
    user_avg = average_rating(reviews)
    model_score = round(statistics.mean(score_from_text(review.content) for review in reviews), 2)
    overall = round(user_avg * 0.65 + model_score * 0.35, 2) if user_avg else model_score
    aspect_scores, aspect_analysis, strengths, weaknesses = summarize_aspects_heuristic(reviews)

    summary = (
        f"{car_model}共分析{len(reviews)}条评论。用户平均评分为{user_avg:.2f}/5，"
        f"文本分析分为{model_score:.2f}/5，综合评分为{overall:.2f}/5。"
    )
    purchase_advice = (
        "整体口碑较强，适合优先试驾并重点确认个人最在意的短板。"
        if overall >= 4.0
        else "建议围绕最关注的维度做深度试驾，再决定是否购买。"
    )

    return AnalysisResult(
        sheet_name=sheet_name,
        car_model=car_model,
        review_count=len(reviews),
        user_avg_rating=user_avg,
        model_score=model_score,
        overall_score=overall,
        summary=summary,
        strengths=strengths,
        weaknesses=weaknesses,
        purchase_advice=purchase_advice,
        aspect_scores=aspect_scores,
        aspect_analysis=aspect_analysis,
        method="heuristic",
    )


def build_llm_payload(reviews: List[Review], max_comments: int) -> List[Dict[str, str]]:
    payload = []
    for index, review in enumerate(reviews[:max_comments], start=1):
        payload.append(
            {
                "index": str(index),
                "time": review.review_time,
                "car_model": review.car_model,
                "rating": "" if review.rating is None else f"{review.rating:.2f}",
                "content": review.content[:1200],
            }
        )
    return payload


def extract_json_from_text(text: str) -> Dict:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        text = text[start : end + 1]
    return json.loads(text)


def llm_analysis(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    car_model = next((review.car_model for review in reviews if review.car_model), sheet_name)
    user_avg = average_rating(reviews)
    client = OpenAI(api_key=API_KEY, base_url=API_BASE)

    prompt = {
        "car_model": car_model,
        "review_count": len(reviews),
        "user_average_rating": user_avg,
        "task": (
            "请你作为汽车评论分析师，基于用户购后评论输出JSON。"
            "需要给出summary、model_score(1-5)、strengths、weaknesses、purchase_advice、"
            "aspect_scores、aspect_analysis。最终评分体系都按5分制输出。"
        ),
        "aspects": list(ASPECTS.keys()),
        "comments": build_llm_payload(reviews, MAX_COMMENTS_FOR_LLM),
        "output_schema": {
            "summary": "string",
            "model_score": "number",
            "strengths": ["string"],
            "weaknesses": ["string"],
            "purchase_advice": "string",
            "aspect_scores": {aspect: "number" for aspect in ASPECTS},
            "aspect_analysis": {aspect: "string" for aspect in ASPECTS},
        },
    }

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": "请只返回合法JSON，不要输出代码块，不要解释。"},
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        temperature=0.2,
    )
    data = extract_json_from_text(response.choices[0].message.content.strip())

    model_score = round(float(data.get("model_score", user_avg or 3.5)), 2)
    overall = round(user_avg * 0.65 + model_score * 0.35, 2) if user_avg else model_score

    aspect_scores = {}
    for aspect in ASPECTS:
        try:
            aspect_scores[aspect] = round(float(data.get("aspect_scores", {}).get(aspect, 0)), 2)
        except Exception:
            aspect_scores[aspect] = 0.0

    aspect_analysis = {aspect: clean_text(data.get("aspect_analysis", {}).get(aspect, "")) for aspect in ASPECTS}

    return AnalysisResult(
        sheet_name=sheet_name,
        car_model=car_model,
        review_count=len(reviews),
        user_avg_rating=user_avg,
        model_score=model_score,
        overall_score=overall,
        summary=clean_text(data.get("summary", "")),
        strengths=[clean_text(item) for item in data.get("strengths", [])][:5],
        weaknesses=[clean_text(item) for item in data.get("weaknesses", [])][:5],
        purchase_advice=clean_text(data.get("purchase_advice", "")),
        aspect_scores=aspect_scores,
        aspect_analysis=aspect_analysis,
        method=f"llm:{MODEL_NAME}",
    )


def analyze_sheet(sheet_name: str, reviews: List[Review]) -> AnalysisResult:
    if ANALYSIS_MODE == "heuristic":
        return heuristic_analysis(sheet_name, reviews)
    if ANALYSIS_MODE == "llm":
        return llm_analysis(sheet_name, reviews)
    if API_KEY and API_KEY != "PUT_YOUR_API_KEY_HERE":
        try:
            return llm_analysis(sheet_name, reviews)
        except Exception:
            return heuristic_analysis(sheet_name, reviews)
    return heuristic_analysis(sheet_name, reviews)


def render_markdown(results: List[AnalysisResult]) -> str:
    lines: List[str] = ["# 汽车评论分析报告", ""]

    lines.append("## 总览")
    lines.append("")
    for index, item in enumerate(sorted(results, key=lambda x: x.overall_score, reverse=True), start=1):
        lines.append(
            f"{index}. {item.car_model}: 综合 {item.overall_score:.2f}/5，"
            f"用户评分均值 {item.user_avg_rating:.2f}/5，模型分析分 {item.model_score:.2f}/5，评论数 {item.review_count}"
        )
    lines.append("")

    for item in results:
        lines.append(f"## {item.car_model}")
        lines.append("")
        lines.append(f"- 分析方式: `{item.method}`")
        lines.append(f"- 评论数: `{item.review_count}`")
        lines.append(f"- 用户平均评分: `{item.user_avg_rating:.2f}/5`")
        lines.append(f"- 模型分析分: `{item.model_score:.2f}/5`")
        lines.append(f"- 综合评分: `{item.overall_score:.2f}/5`")
        lines.append("")
        lines.append("### 模型分析结果")
        lines.append(item.summary)
        lines.append("")
        lines.append("### 主要优点")
        for strength in item.strengths:
            lines.append(f"- {strength}")
        lines.append("")
        lines.append("### 主要短板")
        for weakness in item.weaknesses:
            lines.append(f"- {weakness}")
        lines.append("")
        lines.append("### 用户各方面体验分析")
        for aspect in ASPECTS:
            lines.append(f"- {aspect}: {item.aspect_scores.get(aspect, 0.0):.2f}/5。{item.aspect_analysis.get(aspect, '')}")
        lines.append("")
        lines.append("### 建议")
        lines.append(item.purchase_advice)
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "车型分析"
    return cleaned[:31]


def write_summary_excel(path: str, results: List[AnalysisResult]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "分析汇总"
    summary.append(["车型", "评论数", "用户平均评分", "模型分析分", "综合评分(5分制)", "分析方式", "建议"])

    for item in sorted(results, key=lambda x: x.overall_score, reverse=True):
        summary.append(
            [
                item.car_model,
                item.review_count,
                item.user_avg_rating,
                item.model_score,
                item.overall_score,
                item.method,
                item.purchase_advice,
            ]
        )

    for item in results:
        sheet = workbook.create_sheet(safe_sheet_name(item.car_model))
        sheet.append(["维度", "得分(5分制)", "分析"])
        for aspect in ASPECTS:
            sheet.append([aspect, item.aspect_scores.get(aspect, 0.0), item.aspect_analysis.get(aspect, "")])

    workbook.save(path)


def write_text(path: str, text: str) -> None:
    Path(path).write_text(text, encoding="utf-8")


def main() -> None:
    if not Path(INPUT_EXCEL).exists():
        raise SystemExit(f"Input Excel not found: {INPUT_EXCEL}")

    reviews_by_sheet = load_reviews_by_sheet(INPUT_EXCEL)
    if not reviews_by_sheet:
        raise SystemExit("No review data found in the input Excel.")

    results: List[AnalysisResult] = []
    for sheet_name, reviews in reviews_by_sheet.items():
        results.append(analyze_sheet(sheet_name, reviews))

    markdown = render_markdown(results)
    write_text(OUTPUT_MARKDOWN, markdown)
    write_summary_excel(OUTPUT_SUMMARY_EXCEL, results)

    print(f"input={INPUT_EXCEL}")
    print(f"sheets={len(reviews_by_sheet)}")
    print(f"report={OUTPUT_MARKDOWN}")
    print(f"summary_excel={OUTPUT_SUMMARY_EXCEL}")


if __name__ == "__main__":
    main()
